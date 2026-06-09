import json
import re
import sys
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "db.json"


def normalize(value):
    return str(value or "").strip()


def normalize_curp(value):
    return re.sub(r"\s+", "", normalize(value).upper())


def is_active(value):
    return normalize(value).lower() in {"alta", "activo", "activa", "si", "sí", "true", "1"}


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Uso: python scripts/import_users_from_xlsx.py /ruta/catalogo.xlsx")

    workbook_path = Path(sys.argv[1])
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize(cell.value).lower() for cell in sheet[1]]

    def col(name):
        return headers.index(name.lower())

    idx_clave = col("clave")
    idx_nombre = col("nombre completo")
    idx_departamento = col("departamento")
    idx_estatus = col("estatus")
    idx_curp = col("curp")

    users = []
    seen_curps = set()
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        clave = normalize(row[idx_clave])
        nombre = normalize(row[idx_nombre])
        curp = normalize_curp(row[idx_curp])
        area = normalize(row[idx_departamento]) or "SIN AREA"

        if not clave or not nombre or not curp:
            skipped += 1
            continue

        if curp in seen_curps:
            skipped += 1
            continue

        seen_curps.add(curp)
        users.append(
            {
                "id": f"u-{curp}",
                "nombre": nombre,
                "curp": curp,
                "area": area,
                "activo": is_active(row[idx_estatus]),
            }
        )

    users.sort(key=lambda user: (not user["activo"], user["nombre"]))

    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    valid_user_ids = {user["id"] for user in users}
    db["usuarios"] = users
    db["predicciones"] = [
        prediction for prediction in db.get("predicciones", []) if prediction.get("id_usuario") in valid_user_ids
    ]
    DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    active_count = sum(1 for user in users if user["activo"])
    print(f"Usuarios cargados: {len(users)}")
    print(f"Usuarios activos: {active_count}")
    print(f"Usuarios omitidos: {skipped}")
    print(f"Predicciones conservadas: {len(db['predicciones'])}")
    print(f"Base actualizada: {DB_PATH}")


if __name__ == "__main__":
    main()
