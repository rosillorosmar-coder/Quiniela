import json
import sys
from datetime import datetime, time
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DB_PATH = ROOT / "data" / "db.json"


def parse_time(value):
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    text = str(value).strip()
    if len(text) == 5:
        return text
    return datetime.strptime(text, "%H:%M:%S").strftime("%H:%M")


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Uso: python scripts/import_matches_from_xlsx.py /ruta/calendario.xlsx")

    workbook_path = Path(sys.argv[1])
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook["Partidos"] if "Partidos" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    matches = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        match_number, group, _round, date_value, time_value, home, away, venue = row[:8]
        if match_number is None or not home or not away:
            continue

        if not isinstance(date_value, datetime):
            date_value = datetime.strptime(str(date_value), "%Y-%m-%d")

        match_time = parse_time(time_value)
        date_part = date_value.strftime("%Y-%m-%d")

        matches.append(
            {
                "id": f"m-{int(match_number):03d}",
                "grupo": str(group).strip(),
                "jornada": int(_round) if _round is not None else None,
                "equipo_local": str(home).strip(),
                "equipo_visitante": str(away).strip(),
                "fecha_hora": f"{date_part}T{match_time}:00-06:00",
                "sede": str(venue).strip() if venue else "",
                "goles_local_real": None,
                "goles_visitante_real": None,
                "estatus": "programado",
            }
        )

    matches.sort(key=lambda item: (item["fecha_hora"], item["grupo"], item["id"]))

    db = json.loads(DB_PATH.read_text(encoding="utf-8"))
    db["partidos"] = matches
    db["predicciones"] = []
    DB_PATH.write_text(json.dumps(db, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"Partidos cargados: {len(matches)}")
    print(f"Base actualizada: {DB_PATH}")


if __name__ == "__main__":
    main()
