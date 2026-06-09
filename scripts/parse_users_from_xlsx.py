import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl


def normalize(value):
    return str(value or "").strip()


def normalize_curp(value):
    return re.sub(r"\s+", "", normalize(value).upper())


def normalize_header(value):
    text = normalize(value).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", text)


def is_active(value):
    return normalize(value).lower() in {"alta", "activo", "activa", "si", "sí", "true", "1"}


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Uso: python scripts/parse_users_from_xlsx.py /ruta/catalogo.xlsx")

    workbook_path = Path(sys.argv[1])
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_header(cell.value) for cell in sheet[1]]

    def find_col(*names):
        normalized = [normalize_header(name) for name in names]
        for name in normalized:
            if name in headers:
                return headers.index(name)
        return None

    idx_nombre = find_col("nombre completo", "nombre")
    idx_curp = find_col("curp")
    idx_area = find_col("departamento", "area", "área")
    idx_estatus = find_col("estatus", "activo")

    if idx_nombre is None or idx_curp is None:
        raise SystemExit("El Excel debe incluir columnas de nombre y CURP.")

    users = []
    seen_curps = set()
    skipped = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nombre = normalize(row[idx_nombre] if idx_nombre < len(row) else "")
        curp = normalize_curp(row[idx_curp] if idx_curp < len(row) else "")
        area = normalize(row[idx_area] if idx_area is not None and idx_area < len(row) else "") or "SIN AREA"
        active_value = row[idx_estatus] if idx_estatus is not None and idx_estatus < len(row) else "alta"

        if not nombre or len(curp) != 18:
            skipped += 1
            continue

        if curp in seen_curps:
            skipped += 1
            continue

        seen_curps.add(curp)
        users.append(
            {
                "id": f"u-{curp}",
                "nombre": nombre.upper(),
                "curp": curp,
                "area": area.upper(),
                "activo": is_active(active_value),
            }
        )

    users.sort(key=lambda user: (not user["activo"], user["nombre"]))
    sys.stdout.write(json.dumps({"users": users, "skipped": skipped}, ensure_ascii=False))


if __name__ == "__main__":
    main()
